import requests
import json
import openpyxl

def get_channel_videos(channel_link):
    # Use the YouTube Data API to get information about the channel's videos
    api_key = input("YOUR API KEY: ")
    url = f"https://www.googleapis.com/youtube/v3/search?key={api_key}&channelId={channel_link}&part=snippet,id&order=date&maxResults=50"
    response = requests.get(url)
    data = json.loads(response.text)

    # new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # headers 
    sheet.cell(row=1, column=1).value = "Video Title"
    sheet.cell(row=1, column=2).value = "Video URL"

    
    row = 2
    for item in data["items"]:
        video_title = item["snippet"]["title"]
        video_url = f"https://www.youtube.com/watch?v={item['id']['videoId']}"
        sheet.cell(row=row, column=1).value = video_title
        sheet.cell(row=row, column=2).value = video_url
        row += 1

    
    workbook.save("channel_videos.xlsx")


if __name__ == '__main__':
    channel_link = input('Channel Link: ')
    get_channel_videos(channel_link)